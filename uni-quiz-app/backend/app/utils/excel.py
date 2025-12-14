"""
Excel export utilities for quiz results.
"""
from io import BytesIO
from datetime import datetime
from typing import List, Any
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter


def create_quiz_results_excel(
    quiz_title: str,
    class_name: str,
    results: List[dict]
) -> BytesIO:
    """
    Create an Excel file with quiz results.
    
    Args:
        quiz_title: Title of the quiz
        class_name: Name of the class
        results: List of result dicts containing student info and scores
        
    Returns:
        BytesIO buffer containing the Excel file
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Quiz Results"
    
    # Styles
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")
    
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    pass_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    fail_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    
    # Title section
    ws.merge_cells('A1:G1')
    ws['A1'] = f"Quiz Results: {quiz_title}"
    ws['A1'].font = Font(bold=True, size=14)
    ws['A1'].alignment = Alignment(horizontal="center")
    
    ws.merge_cells('A2:G2')
    ws['A2'] = f"Class: {class_name}"
    ws['A2'].alignment = Alignment(horizontal="center")
    
    ws.merge_cells('A3:G3')
    ws['A3'] = f"Exported: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
    ws['A3'].alignment = Alignment(horizontal="center")
    
    # Headers (row 5)
    headers = [
        "Rank",
        "Student Name",
        "Telegram Username",
        "Student ID",
        "Score (%)",
        "Points",
        "Status",
        "Start Time",
        "Submission Time",
        "Time Spent"
    ]
    
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=5, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = border
    
    # Data rows
    for idx, result in enumerate(results, 1):
        row = idx + 5
        
        # Rank
        ws.cell(row=row, column=1, value=idx).border = border
        
        # Student Name
        ws.cell(row=row, column=2, value=result.get('student_name', 'Unknown')).border = border
        
        # Telegram Username
        ws.cell(row=row, column=3, value=result.get('telegram_username', '-')).border = border
        
        # Student ID
        ws.cell(row=row, column=4, value=result.get('student_id', '-')).border = border
        
        # Score (%)
        score_cell = ws.cell(row=row, column=5, value=result.get('score', 0))
        score_cell.border = border
        score_cell.alignment = Alignment(horizontal="center")
        
        # Points
        points = f"{result.get('earned_points', 0)}/{result.get('total_points', 0)}"
        ws.cell(row=row, column=6, value=points).border = border
        
        # Status
        passed = result.get('passed', False)
        status_cell = ws.cell(row=row, column=7, value="PASSED" if passed else "FAILED")
        status_cell.border = border
        status_cell.fill = pass_fill if passed else fail_fill
        status_cell.alignment = Alignment(horizontal="center")
        
        # Start Time
        start_time = result.get('started_at')
        if start_time:
            if isinstance(start_time, str):
                ws.cell(row=row, column=8, value=start_time).border = border
            else:
                ws.cell(row=row, column=8, value=start_time.strftime('%Y-%m-%d %H:%M')).border = border
        else:
            ws.cell(row=row, column=8, value='-').border = border
        
        # Submission Time
        submit_time = result.get('submitted_at')
        if submit_time:
            if isinstance(submit_time, str):
                ws.cell(row=row, column=9, value=submit_time).border = border
            else:
                ws.cell(row=row, column=9, value=submit_time.strftime('%Y-%m-%d %H:%M')).border = border
        else:
            ws.cell(row=row, column=9, value='-').border = border
        
        # Time Spent
        time_spent = result.get('time_spent_seconds')
        if time_spent:
            minutes = time_spent // 60
            seconds = time_spent % 60
            ws.cell(row=row, column=10, value=f"{minutes}m {seconds}s").border = border
        else:
            ws.cell(row=row, column=10, value='-').border = border
    
    # Summary section
    summary_row = len(results) + 7
    ws.cell(row=summary_row, column=1, value="Summary").font = Font(bold=True)
    
    total_students = len(results)
    passed_count = sum(1 for r in results if r.get('passed', False))
    avg_score = sum(r.get('score', 0) for r in results) / total_students if total_students > 0 else 0
    
    ws.cell(row=summary_row + 1, column=1, value=f"Total Students: {total_students}")
    ws.cell(row=summary_row + 2, column=1, value=f"Passed: {passed_count}")
    ws.cell(row=summary_row + 3, column=1, value=f"Failed: {total_students - passed_count}")
    ws.cell(row=summary_row + 4, column=1, value=f"Average Score: {avg_score:.2f}%")
    ws.cell(row=summary_row + 5, column=1, value=f"Pass Rate: {(passed_count/total_students*100) if total_students > 0 else 0:.1f}%")
    
    # Adjust column widths
    column_widths = [8, 25, 20, 15, 12, 12, 10, 18, 18, 12]
    for col, width in enumerate(column_widths, 1):
        ws.column_dimensions[get_column_letter(col)].width = width
    
    # Save to BytesIO
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    
    return output


def create_question_bank_excel(questions: List[dict]) -> BytesIO:
    """
    Create an Excel file with question bank for export.
    
    Args:
        questions: List of question dicts
        
    Returns:
        BytesIO buffer containing the Excel file
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Question Bank"
    
    # Headers
    headers = [
        "Question Text",
        "Type",
        "Option A",
        "Option B",
        "Option C",
        "Option D",
        "Correct Answer",
        "Explanation",
        "Points",
        "Difficulty",
        "Tags"
    ]
    
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
    
    # Data rows
    for idx, q in enumerate(questions, 2):
        ws.cell(row=idx, column=1, value=q.get('question_text', ''))
        ws.cell(row=idx, column=2, value=q.get('question_type', 'multiple_choice'))
        
        # Options
        options = q.get('options', [])
        for opt_idx, opt in enumerate(options[:4]):
            if isinstance(opt, dict):
                ws.cell(row=idx, column=3 + opt_idx, value=opt.get('text', ''))
            else:
                ws.cell(row=idx, column=3 + opt_idx, value=str(opt))
        
        ws.cell(row=idx, column=7, value=q.get('correct_answer', ''))
        ws.cell(row=idx, column=8, value=q.get('explanation', ''))
        ws.cell(row=idx, column=9, value=q.get('points', 1))
        ws.cell(row=idx, column=10, value=q.get('difficulty', 'medium'))
        
        tags = q.get('tags', [])
        ws.cell(row=idx, column=11, value=', '.join(tags) if tags else '')
    
    # Adjust column widths
    ws.column_dimensions['A'].width = 50
    ws.column_dimensions['B'].width = 15
    for col in ['C', 'D', 'E', 'F']:
        ws.column_dimensions[col].width = 20
    ws.column_dimensions['G'].width = 15
    ws.column_dimensions['H'].width = 40
    ws.column_dimensions['I'].width = 10
    ws.column_dimensions['J'].width = 12
    ws.column_dimensions['K'].width = 25
    
    # Save to BytesIO
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    
    return output
